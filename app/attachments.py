from __future__ import annotations

import base64
import io
import re
import zipfile
import xml.etree.ElementTree as ET
from html import unescape
from pathlib import Path

from app.document_text import extract_pdf_text_from_path, truncate_text
from app.i18n import normalize_locale

_OLE2_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
_MSG_MARKERS_ASCII = (b"__substg1.0_", b"IPM.")
_MSG_MARKERS_UTF16 = tuple(marker.decode("ascii").encode("utf-16-le") for marker in _MSG_MARKERS_ASCII)
_XLSX_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}
_PPTX_SUFFIXES = {".pptx", ".pptm"}
_SAFE_IMAGE_MIMES = {"image/png", "image/jpeg", "image/webp", "image/gif"}

_ATTACHMENT_TEXT: dict[str, dict[str, str]] = {
    "zh-CN": {
        "xlsx.missing_dependency": "解析 .xlsx 需要依赖 openpyxl。请执行 `pip install -r requirements.txt` 后重试。",
        "xlsx.header": "[Excel 工作簿解析]",
        "xlsx.empty": "[空表或无可读内容]",
        "xlsx.truncated": "[内容已截断，工作簿内容较大]",
        "pptx.invalid_zip": "[PPTX 解析失败: 文件不是合法 ZIP 容器]",
        "pptx.empty": "[PPTX 解析结果为空: 未找到 slide XML]",
        "pptx.header": "[PowerPoint 文档解析]",
        "pptx.slide_count": "幻灯片数量: {count}",
        "pptx.truncated": "[内容已截断，幻灯片内容较多]",
        "pptx.no_text": "[未提取到文本（可能是纯图片页）]",
        "xml.atom": "[Atom Feed 解析]",
        "xml.rss": "[RSS Feed 解析]",
        "xml.title": "标题: {value}",
        "xml.subtitle": "副标题: {value}",
        "xml.updated": "更新时间: {value}",
        "xml.author": "作者: {value}",
        "xml.description": "描述: {value}",
        "xml.entries": "条目:",
        "xml.time": "时间: {value}",
        "xml.summary": "摘要: {value}",
        "msg.nested_mail": "嵌套邮件: {value}",
        "msg.header": "[Outlook MSG 邮件解析]",
        "msg.class_type": "消息类型: {value}",
        "msg.subject": "主题: {value}",
        "msg.sender": "发件人: {value}",
        "msg.to": "收件人: {value}",
        "msg.cc": "抄送: {value}",
        "msg.date": "时间: {value}",
        "msg.attachments": "附件列表:",
        "msg.body_separator": "--- 正文 ---",
        "msg.no_body": "[未提取到可读正文：该邮件可能仅包含附件、图片或受限富文本内容]",
        "msg.missing_dependency": "解析 .msg 需要依赖 extract-msg。请执行 `pip install -r requirements.txt` 后重试。",
        "doc.unsupported_xls": "[暂不支持 .xls（二进制 Excel）直接解析，请先另存为 .xlsx 后再读取]",
        "doc.unsupported_ppt": "[暂不支持 .ppt（二进制 PowerPoint）直接解析，请先另存为 .pptx 后再读取]",
        "doc.parse_failed": "[文档解析失败: {error}]",
        "image.heic_warning": "HEIC 未本地转码，已原始上传；若网关不支持 HEIC，请先转 JPG/PNG。",
        "image.converted_warning": "检测到非标准图片类型({mime})，已转为 PNG 再发送。",
        "image.unsupported_type": "不支持的图片类型({mime})，且转码失败: {error}",
        "image.empty": "图片内容为空，无法编码为 data URL。",
        "file.empty": "[空文件]",
        "file.text_preview": "[文本预览，文件大小 {size} bytes]\\n{text}",
        "file.binary_preview": "[二进制预览，文件大小 {size} bytes，前 {preview_size} bytes(hex)]\\n{hex_preview}",
    },
    "ja-JP": {
        "xlsx.missing_dependency": ".xlsx の解析には openpyxl が必要です。`pip install -r requirements.txt` を実行してから再試行してください。",
        "xlsx.header": "[Excel ブック解析]",
        "xlsx.empty": "[空シート、または読み取れる内容がありません]",
        "xlsx.truncated": "[内容を切り詰めました。ブックの内容が大きすぎます]",
        "pptx.invalid_zip": "[PPTX 解析失敗: 正しい ZIP コンテナではありません]",
        "pptx.empty": "[PPTX 解析結果なし: slide XML が見つかりません]",
        "pptx.header": "[PowerPoint ドキュメント解析]",
        "pptx.slide_count": "スライド数: {count}",
        "pptx.truncated": "[内容を切り詰めました。スライド内容が多すぎます]",
        "pptx.no_text": "[テキストを抽出できませんでした（画像のみのスライドの可能性があります）]",
        "xml.atom": "[Atom フィード解析]",
        "xml.rss": "[RSS フィード解析]",
        "xml.title": "タイトル: {value}",
        "xml.subtitle": "サブタイトル: {value}",
        "xml.updated": "更新日時: {value}",
        "xml.author": "作成者: {value}",
        "xml.description": "説明: {value}",
        "xml.entries": "項目:",
        "xml.time": "時刻: {value}",
        "xml.summary": "要約: {value}",
        "msg.nested_mail": "ネストされたメール: {value}",
        "msg.header": "[Outlook MSG メール解析]",
        "msg.class_type": "メッセージ種別: {value}",
        "msg.subject": "件名: {value}",
        "msg.sender": "送信者: {value}",
        "msg.to": "宛先: {value}",
        "msg.cc": "CC: {value}",
        "msg.date": "日時: {value}",
        "msg.attachments": "添付一覧:",
        "msg.body_separator": "--- 本文 ---",
        "msg.no_body": "[読み取れる本文を抽出できませんでした。このメールには添付、画像、または制限付きリッチテキストのみが含まれている可能性があります]",
        "msg.missing_dependency": ".msg の解析には extract-msg が必要です。`pip install -r requirements.txt` を実行してから再試行してください。",
        "doc.unsupported_xls": "[.xls（バイナリ Excel）は直接解析できません。.xlsx に保存し直してから読み取ってください]",
        "doc.unsupported_ppt": "[.ppt（バイナリ PowerPoint）は直接解析できません。.pptx に保存し直してから読み取ってください]",
        "doc.parse_failed": "[ドキュメント解析失敗: {error}]",
        "image.heic_warning": "HEIC はローカル変換できなかったため元の形式で送信しました。ゲートウェイが HEIC をサポートしない場合は JPG/PNG に変換してください。",
        "image.converted_warning": "標準外の画像タイプ ({mime}) を検出したため、PNG に変換して送信しました。",
        "image.unsupported_type": "未対応の画像タイプ ({mime}) で、変換にも失敗しました: {error}",
        "image.empty": "画像内容が空のため data URL に変換できません。",
        "file.empty": "[空ファイル]",
        "file.text_preview": "[テキストプレビュー、ファイルサイズ {size} bytes]\\n{text}",
        "file.binary_preview": "[バイナリプレビュー、ファイルサイズ {size} bytes、先頭 {preview_size} bytes(hex)]\\n{hex_preview}",
    },
    "en": {
        "xlsx.missing_dependency": "Parsing .xlsx files requires openpyxl. Run `pip install -r requirements.txt` and try again.",
        "xlsx.header": "[Excel Workbook Parse]",
        "xlsx.empty": "[Empty sheet or no readable content]",
        "xlsx.truncated": "[Content truncated because the workbook is too large]",
        "pptx.invalid_zip": "[PPTX parse failed: file is not a valid ZIP container]",
        "pptx.empty": "[PPTX parse returned no result: slide XML not found]",
        "pptx.header": "[PowerPoint Document Parse]",
        "pptx.slide_count": "Slide count: {count}",
        "pptx.truncated": "[Content truncated because the slide deck is large]",
        "pptx.no_text": "[No text extracted; this slide may contain only images]",
        "xml.atom": "[Atom Feed Parse]",
        "xml.rss": "[RSS Feed Parse]",
        "xml.title": "Title: {value}",
        "xml.subtitle": "Subtitle: {value}",
        "xml.updated": "Updated: {value}",
        "xml.author": "Author: {value}",
        "xml.description": "Description: {value}",
        "xml.entries": "Entries:",
        "xml.time": "Time: {value}",
        "xml.summary": "Summary: {value}",
        "msg.nested_mail": "Nested email: {value}",
        "msg.header": "[Outlook MSG Email Parse]",
        "msg.class_type": "Message type: {value}",
        "msg.subject": "Subject: {value}",
        "msg.sender": "Sender: {value}",
        "msg.to": "To: {value}",
        "msg.cc": "CC: {value}",
        "msg.date": "Date: {value}",
        "msg.attachments": "Attachments:",
        "msg.body_separator": "--- Body ---",
        "msg.no_body": "[No readable body was extracted. The message may contain only attachments, images, or restricted rich text content.]",
        "msg.missing_dependency": "Parsing .msg files requires extract-msg. Run `pip install -r requirements.txt` and try again.",
        "doc.unsupported_xls": "[Binary .xls files are not directly supported. Save the file as .xlsx first, then read it again.]",
        "doc.unsupported_ppt": "[Binary .ppt files are not directly supported. Save the file as .pptx first, then read it again.]",
        "doc.parse_failed": "[Document parse failed: {error}]",
        "image.heic_warning": "HEIC was uploaded without local conversion. If your gateway does not support HEIC, convert it to JPG/PNG first.",
        "image.converted_warning": "A non-standard image type ({mime}) was detected and converted to PNG before sending.",
        "image.unsupported_type": "Unsupported image type ({mime}), and conversion failed: {error}",
        "image.empty": "The image is empty and cannot be encoded as a data URL.",
        "file.empty": "[Empty file]",
        "file.text_preview": "[Text preview, file size {size} bytes]\\n{text}",
        "file.binary_preview": "[Binary preview, file size {size} bytes, first {preview_size} bytes(hex)]\\n{hex_preview}",
    },
}


def _attachment_text(locale: str, key: str, **values: object) -> str:
    normalized = normalize_locale(locale, fallback="zh-CN")
    template = (
        (_ATTACHMENT_TEXT.get(normalized) or {}).get(key)
        or (_ATTACHMENT_TEXT.get("en") or {}).get(key)
        or (_ATTACHMENT_TEXT.get("zh-CN") or {}).get(key)
        or key
    )
    if not values:
        return template
    try:
        return template.format(**values)
    except Exception:
        return template


def _read_plain_text(path: Path, max_chars: int) -> str:
    text = path.read_text(encoding="utf-8", errors="ignore")
    return truncate_text(text, max_chars)


def _extract_pdf(path: Path, max_chars: int) -> str:
    return extract_pdf_text_from_path(path, max_chars=max_chars)


def _extract_docx(path: Path, max_chars: int) -> str:
    from docx import Document  # lazy import

    doc = Document(str(path))
    text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    return truncate_text(text, max_chars)


def _xlsx_cell_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float)):
        try:
            if isinstance(value, float) and value.is_integer():
                return str(int(value))
        except Exception:
            pass
        return str(value)
    if hasattr(value, "isoformat"):
        try:
            return str(value.isoformat())
        except Exception:
            pass
    return str(value).strip()


def looks_like_xlsx_file(path: Path) -> bool:
    try:
        if not zipfile.is_zipfile(path):
            return False
        with zipfile.ZipFile(path, "r") as zf:
            names = set(zf.namelist())
        return "xl/workbook.xml" in names
    except Exception:
        return False


def looks_like_pptx_file(path: Path) -> bool:
    try:
        if not zipfile.is_zipfile(path):
            return False
        with zipfile.ZipFile(path, "r") as zf:
            names = set(zf.namelist())
        if "ppt/presentation.xml" not in names:
            return False
        return any(name.startswith("ppt/slides/slide") and name.endswith(".xml") for name in names)
    except Exception:
        return False


def _extract_xlsx(path: Path, max_chars: int, *, locale: str = "zh-CN") -> str:
    try:
        from openpyxl import load_workbook  # lazy import
    except Exception as exc:
        raise RuntimeError(_attachment_text(locale, "xlsx.missing_dependency")) from exc

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        lines: list[str] = [_attachment_text(locale, "xlsx.header")]
        total_chars = len(lines[0])
        truncated = False
        for sheet in wb.worksheets:
            title = (sheet.title or "").strip() or "Sheet"
            sheet_header = f"\n--- Sheet: {title} ---"
            lines.append(sheet_header)
            total_chars += len(sheet_header)
            if total_chars >= max_chars:
                truncated = True
                break

            sheet_rows = 0
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                cells = [_xlsx_cell_to_text(cell) for cell in row]
                while cells and not cells[-1]:
                    cells.pop()
                if not cells or not any(cells):
                    continue

                row_line = f"{row_idx}: " + " | ".join(cells)
                lines.append(row_line)
                total_chars += len(row_line)
                sheet_rows += 1
                if total_chars >= max_chars:
                    truncated = True
                    break

            if sheet_rows == 0:
                empty_line = _attachment_text(locale, "xlsx.empty")
                lines.append(empty_line)
                total_chars += len(empty_line)
            if truncated:
                break

        if truncated:
            lines.append(f"\n{_attachment_text(locale, 'xlsx.truncated')}")
        return truncate_text("\n".join(lines), max_chars)
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _ppt_xml_to_lines(raw_xml: bytes, per_slide_limit: int = 40) -> list[str]:
    try:
        root = ET.fromstring(raw_xml)
    except Exception:
        return []
    lines: list[str] = []
    current: list[str] = []
    for node in root.iter():
        if _xml_local_name(node.tag) != "t":
            continue
        text = " ".join(str(node.text or "").split()).strip()
        if not text:
            continue
        current.append(text)
        # Keep nearby text runs together as one line.
        if len(current) >= 12:
            merged = " ".join(current).strip()
            if merged:
                lines.append(merged)
            current = []
        if len(lines) >= per_slide_limit:
            break
    if current and len(lines) < per_slide_limit:
        merged = " ".join(current).strip()
        if merged:
            lines.append(merged)
    # Deduplicate preserving order.
    out: list[str] = []
    seen: set[str] = set()
    for line in lines:
        key = line.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(line)
    return out


def _extract_pptx(path: Path, max_chars: int, *, locale: str = "zh-CN") -> str:
    if not zipfile.is_zipfile(path):
        return _attachment_text(locale, "pptx.invalid_zip")

    with zipfile.ZipFile(path, "r") as zf:
        names = zf.namelist()
        slide_names = [
            name
            for name in names
            if name.startswith("ppt/slides/slide") and name.endswith(".xml")
        ]
        if not slide_names:
            return _attachment_text(locale, "pptx.empty")

        def sort_key(name: str) -> tuple[int, str]:
            m = re.search(r"slide(\d+)\.xml$", name)
            return (int(m.group(1)) if m else 10**9, name)

        slide_names.sort(key=sort_key)

        lines: list[str] = [_attachment_text(locale, "pptx.header")]
        lines.append(_attachment_text(locale, "pptx.slide_count", count=len(slide_names)))
        total = sum(len(x) for x in lines) + 1

        for idx, slide_name in enumerate(slide_names, start=1):
            try:
                raw_xml = zf.read(slide_name)
            except Exception:
                raw_xml = b""
            slide_lines = _ppt_xml_to_lines(raw_xml, per_slide_limit=36)
            header = f"\n--- Slide {idx} ---"
            if total + len(header) >= max_chars:
                lines.append(f"\n{_attachment_text(locale, 'pptx.truncated')}")
                break
            lines.append(header)
            total += len(header)

            if not slide_lines:
                placeholder = _attachment_text(locale, "pptx.no_text")
                if total + len(placeholder) >= max_chars:
                    lines.append(f"\n{_attachment_text(locale, 'pptx.truncated')}")
                    break
                lines.append(placeholder)
                total += len(placeholder)
                continue

            for line in slide_lines:
                entry = f"- {line}"
                if total + len(entry) + 1 >= max_chars:
                    lines.append(f"\n{_attachment_text(locale, 'pptx.truncated')}")
                    break
                lines.append(entry)
                total += len(entry) + 1
            else:
                continue
            break

    return truncate_text("\n".join(lines), max_chars)


def _html_to_text(html: str) -> str:
    raw = html or ""
    raw = re.sub(r"(?is)<(script|style).*?>.*?</\1>", " ", raw)
    raw = re.sub(r"(?i)<br\\s*/?>", "\n", raw)
    raw = re.sub(r"(?i)</(p|div|li|tr|h1|h2|h3|h4|h5|h6|section|article)>", "\n", raw)
    raw = re.sub(r"(?s)<[^>]+>", " ", raw)
    raw = unescape(raw)
    lines: list[str] = []
    for line in raw.splitlines():
        normalized = re.sub(r"\s+", " ", line).strip()
        if normalized:
            lines.append(normalized)
    return "\n".join(lines)


def _xml_local_name(tag: str) -> str:
    raw = str(tag or "").strip()
    if "}" in raw:
        raw = raw.rsplit("}", 1)[-1]
    return raw.lower()


def _find_first_child_text(node: ET.Element, *names: str) -> str:
    wanted = {name.lower() for name in names if name}
    for child in list(node):
        if _xml_local_name(child.tag) not in wanted:
            continue
        text = "".join(child.itertext()).strip()
        if text:
            return text
    return ""


def _extract_xml_feed(path: Path, max_chars: int, *, locale: str = "zh-CN") -> str:
    raw_text = path.read_text(encoding="utf-8", errors="ignore")
    if not raw_text.strip():
        return ""

    try:
        root = ET.fromstring(raw_text)
    except Exception:
        return truncate_text(raw_text, max_chars)

    lines: list[str] = []
    root_name = _xml_local_name(root.tag)

    if root_name == "feed":
        lines.append(_attachment_text(locale, "xml.atom"))
        title = _find_first_child_text(root, "title")
        subtitle = _find_first_child_text(root, "subtitle", "tagline")
        updated = _find_first_child_text(root, "updated")
        feed_id = _find_first_child_text(root, "id")
        author_name = ""
        for child in list(root):
            if _xml_local_name(child.tag) == "author":
                author_name = _find_first_child_text(child, "name") or "".join(child.itertext()).strip()
                if author_name:
                    break

        if title:
            lines.append(_attachment_text(locale, "xml.title", value=title))
        if subtitle:
            lines.append(_attachment_text(locale, "xml.subtitle", value=subtitle))
        if updated:
            lines.append(_attachment_text(locale, "xml.updated", value=updated))
        if author_name:
            lines.append(_attachment_text(locale, "xml.author", value=author_name))
        if feed_id:
            lines.append(f"Feed ID: {feed_id}")

        entries = [child for child in list(root) if _xml_local_name(child.tag) == "entry"]
        if entries:
            lines.append(_attachment_text(locale, "xml.entries"))
        for idx, entry in enumerate(entries, start=1):
            entry_title = _find_first_child_text(entry, "title") or f"entry_{idx}"
            entry_updated = _find_first_child_text(entry, "updated", "published")
            entry_summary = _find_first_child_text(entry, "summary", "content")
            lines.append(f"{idx}. {entry_title}")
            if entry_updated:
                lines.append(f"   {_attachment_text(locale, 'xml.updated', value=entry_updated)}")
            if entry_summary:
                entry_summary_clean = re.sub(r"\s+", " ", entry_summary).strip()
                lines.append(f"   {_attachment_text(locale, 'xml.summary', value=entry_summary_clean)}")
    elif root_name == "rss":
        lines.append(_attachment_text(locale, "xml.rss"))
        channel = next((child for child in list(root) if _xml_local_name(child.tag) == "channel"), None)
        channel_node = channel or root
        title = _find_first_child_text(channel_node, "title")
        description = _find_first_child_text(channel_node, "description")
        updated = _find_first_child_text(channel_node, "lastbuilddate", "pubdate")
        if title:
            lines.append(_attachment_text(locale, "xml.title", value=title))
        if description:
            lines.append(_attachment_text(locale, "xml.description", value=description))
        if updated:
            lines.append(_attachment_text(locale, "xml.updated", value=updated))

        items = [child for child in list(channel_node) if _xml_local_name(child.tag) == "item"]
        if items:
            lines.append(_attachment_text(locale, "xml.entries"))
        for idx, item in enumerate(items, start=1):
            item_title = _find_first_child_text(item, "title") or f"item_{idx}"
            item_date = _find_first_child_text(item, "pubdate")
            item_desc = _find_first_child_text(item, "description", "summary")
            lines.append(f"{idx}. {item_title}")
            if item_date:
                lines.append(f"   {_attachment_text(locale, 'xml.time', value=item_date)}")
            if item_desc:
                item_desc_clean = re.sub(r"\s+", " ", item_desc).strip()
                lines.append(f"   {_attachment_text(locale, 'xml.summary', value=item_desc_clean)}")
    else:
        return truncate_text(raw_text, max_chars)

    return truncate_text("\n".join(lines).strip(), max_chars)


def _decode_bytes_best_effort(raw: bytes) -> str:
    if not raw:
        return ""
    for encoding in ("utf-8", "utf-16-le", "utf-16-be", "latin-1"):
        try:
            out = raw.decode(encoding, errors="ignore")
        except Exception:
            continue
        if out.strip():
            return out
    return raw.decode("utf-8", errors="ignore")


def _looks_binaryish_text(text: str) -> bool:
    if not text:
        return False
    sample = text[:4096]
    if not sample:
        return False

    bad = 0
    for ch in sample:
        code = ord(ch)
        if code == 0:
            bad += 3
        elif code < 32 and ch not in "\n\r\t":
            bad += 1

    ratio = bad / max(1, len(sample))
    return ratio >= 0.02


def looks_like_outlook_msg_bytes(raw: bytes) -> bool:
    if not raw or not raw.startswith(_OLE2_MAGIC):
        return False
    head = raw[: max(4096, min(len(raw), 512 * 1024))]
    if any(marker in head for marker in _MSG_MARKERS_ASCII):
        return True
    if any(marker in head for marker in _MSG_MARKERS_UTF16):
        return True
    return False


def looks_like_outlook_msg_file(path: Path) -> bool:
    try:
        with path.open("rb") as fp:
            head = fp.read(512 * 1024)
    except Exception:
        return False
    return looks_like_outlook_msg_bytes(head)


def _extract_msg_body(msg: object) -> str:
    body = ""
    try:
        plain = getattr(msg, "body", None)
        if isinstance(plain, str):
            body = plain.strip()
        elif isinstance(plain, (bytes, bytearray)):
            body = _decode_bytes_best_effort(bytes(plain)).strip()
    except Exception:
        body = ""
    if body and not _looks_binaryish_text(body):
        return body

    try:
        html_body = getattr(msg, "htmlBody", None)
        if isinstance(html_body, (bytes, bytearray)):
            html_body = _decode_bytes_best_effort(bytes(html_body))
        if isinstance(html_body, str) and html_body.strip():
            html_text = _html_to_text(html_body).strip()
            if html_text and not _looks_binaryish_text(html_text):
                return html_text
    except Exception:
        pass

    try:
        rtf_body = getattr(msg, "rtfBody", None)
        deencap = getattr(msg, "deencapsulateBody", None)
        if rtf_body and callable(deencap):
            try:
                from extract_msg.enums import DeencapType  # lazy import

                plain_rtf = deencap(rtf_body, DeencapType.PLAIN)
            except Exception:
                plain_rtf = None
            if isinstance(plain_rtf, (bytes, bytearray)):
                plain_rtf = _decode_bytes_best_effort(bytes(plain_rtf))
            if isinstance(plain_rtf, str):
                plain_rtf = plain_rtf.strip()
                if plain_rtf and not _looks_binaryish_text(plain_rtf):
                    return plain_rtf
    except Exception:
        pass

    return ""


def _format_msg_attachment_line(att: object, idx: int, *, locale: str = "zh-CN") -> str:
    name = (
        (getattr(att, "longFilename", None) or "")
        or (getattr(att, "filename", None) or "")
        or (getattr(att, "name", None) or "")
        or f"attachment_{idx}"
    )
    extras: list[str] = []

    att_type = str(getattr(att, "type", "") or "").strip()
    if att_type:
        extras.append(att_type.split(".")[-1].lower())

    mime = (getattr(att, "mimetype", None) or "").strip()
    if mime:
        extras.append(mime)

    data = None
    try:
        data = getattr(att, "data", None)
    except Exception:
        data = None

    if isinstance(data, (bytes, bytearray)):
        extras.append(f"{len(data)} bytes")
    else:
        nested_subject = (getattr(data, "subject", None) or "").strip() if data is not None else ""
        if nested_subject:
            extras.append(_attachment_text(locale, "msg.nested_mail", value=nested_subject))

    if extras:
        return f"- {name} ({', '.join(extras)})"
    return f"- {name}"


def _extract_msg_attachment_payload(att: object, idx: int) -> dict[str, object]:
    name = (
        (getattr(att, "longFilename", None) or "").strip()
        or (getattr(att, "filename", None) or "").strip()
        or f"attachment_{idx}"
    )
    mime_hint = str(getattr(att, "mimetype", None) or "").strip()

    size: int | None = None
    try:
        data = getattr(att, "data", None)
    except Exception:
        data = None
    if isinstance(data, (bytes, bytearray)):
        size = len(data)

    return {
        "name": name,
        "size": size,
        "mime_hint": mime_hint,
    }


def _render_outlook_msg_content(
    *,
    class_type: str,
    subject: str,
    sender: str,
    to: str,
    cc: str,
    date: str,
    body: str,
    attachment_lines: list[str],
    max_chars: int,
    locale: str = "zh-CN",
) -> str:
    sections: list[str] = [_attachment_text(locale, "msg.header")]
    if class_type:
        sections.append(_attachment_text(locale, "msg.class_type", value=class_type))
    if subject:
        sections.append(_attachment_text(locale, "msg.subject", value=subject))
    if sender:
        sections.append(_attachment_text(locale, "msg.sender", value=sender))
    if to:
        sections.append(_attachment_text(locale, "msg.to", value=to))
    if cc:
        sections.append(_attachment_text(locale, "msg.cc", value=cc))
    if date:
        sections.append(_attachment_text(locale, "msg.date", value=date))
    if attachment_lines:
        sections.append(_attachment_text(locale, "msg.attachments"))
        sections.extend(attachment_lines)
    body_separator = _attachment_text(locale, "msg.body_separator")
    if body:
        sections.append(f"\n{body_separator}\n")
        sections.append(body)
    else:
        sections.append(f"\n{body_separator}\n")
        sections.append(_attachment_text(locale, "msg.no_body"))

    return truncate_text("\n".join(sections).strip(), max_chars)


def _extract_outlook_msg_payload(path: Path, max_chars: int, *, locale: str = "zh-CN") -> dict[str, object]:
    try:
        import extract_msg  # lazy import
    except Exception as exc:
        raise RuntimeError(_attachment_text(locale, "msg.missing_dependency")) from exc

    msg = extract_msg.openMsg(str(path), strict=False, delayAttachments=False)
    try:
        subject = (msg.subject or "").strip()
        sender = (msg.sender or "").strip()
        to = (msg.to or "").strip()
        cc = (msg.cc or "").strip()
        date = str(msg.date or "").strip()
        class_type = str(getattr(msg, "classType", "") or "").strip()
        body = _extract_msg_body(msg)

        attachment_list: list[dict[str, object]] = []
        attachment_lines: list[str] = []
        for idx, att in enumerate(getattr(msg, "attachments", []) or [], start=1):
            attachment_list.append(_extract_msg_attachment_payload(att, idx))
            attachment_lines.append(_format_msg_attachment_line(att, idx, locale=locale))

        email_meta = {
            "subject": subject,
            "sender": sender,
            "to": to,
            "cc": cc,
            "date": date,
            "class_type": class_type,
        }
        content = _render_outlook_msg_content(
            class_type=class_type,
            subject=subject,
            sender=sender,
            to=to,
            cc=cc,
            date=date,
            body=body,
            attachment_lines=attachment_lines,
            max_chars=max_chars,
            locale=locale,
        )
        return {
            "content": content,
            "email_meta": email_meta,
            "attachment_list": attachment_list,
        }
    finally:
        close = getattr(msg, "close", None)
        if callable(close):
            try:
                close()
            except Exception:
                pass


def extract_outlook_msg_payload(path: str, max_chars: int, *, locale: str = "zh-CN") -> dict[str, object] | None:
    try:
        return _extract_outlook_msg_payload(Path(path), max_chars, locale=locale)
    except Exception:
        return None


def _extract_outlook_msg(path: Path, max_chars: int, *, locale: str = "zh-CN") -> str:
    payload = _extract_outlook_msg_payload(path, max_chars, locale=locale)
    return str(payload.get("content") or "")


def extract_document_text(path: str, max_chars: int, *, locale: str = "zh-CN") -> str | None:
    file_path = Path(path)
    suffix = file_path.suffix.lower()

    plain_suffixes = {
        ".atom",
        ".txt",
        ".md",
        ".csv",
        ".json",
        ".log",
        ".py",
        ".js",
        ".ts",
        ".tsx",
        ".html",
        ".css",
        ".yaml",
        ".yml",
        ".xml",
        ".rss",
    }

    try:
        if suffix in {".atom", ".rss", ".xml"}:
            return _extract_xml_feed(file_path, max_chars, locale=locale)
        if suffix in plain_suffixes:
            return _read_plain_text(file_path, max_chars)
        if suffix == ".pdf":
            return _extract_pdf(file_path, max_chars)
        if suffix == ".docx":
            return _extract_docx(file_path, max_chars)
        if suffix in _XLSX_SUFFIXES:
            return _extract_xlsx(file_path, max_chars, locale=locale)
        if suffix in _PPTX_SUFFIXES:
            return _extract_pptx(file_path, max_chars, locale=locale)
        if suffix == ".xls":
            return _attachment_text(locale, "doc.unsupported_xls")
        if suffix == ".ppt":
            return _attachment_text(locale, "doc.unsupported_ppt")
        if suffix in {".zip", ".bin"} and looks_like_xlsx_file(file_path):
            return _extract_xlsx(file_path, max_chars, locale=locale)
        if suffix in {".zip", ".bin"} and looks_like_pptx_file(file_path):
            return _extract_pptx(file_path, max_chars, locale=locale)
        if suffix == ".msg" or looks_like_outlook_msg_file(file_path):
            payload = extract_outlook_msg_payload(str(file_path), max_chars, locale=locale)
            return str((payload or {}).get("content") or "")
    except Exception as exc:
        return _attachment_text(locale, "doc.parse_failed", error=exc)

    return None


def _heic_to_jpeg_bytes(path: Path) -> bytes:
    try:
        from PIL import Image
        from pillow_heif import register_heif_opener

        register_heif_opener()
        image = Image.open(path)
        rgb = image.convert("RGB")
        buffer = io.BytesIO()
        rgb.save(buffer, format="JPEG", quality=92)
        return buffer.getvalue()
    except Exception as exc:
        raise RuntimeError(
            "HEIC/HEIF conversion requires pillow-heif. Please install dependencies from requirements.txt."
        ) from exc


def _normalize_image_mime(mime: str, suffix: str) -> str:
    normalized = str(mime or "").strip().lower()
    if ";" in normalized:
        normalized = normalized.split(";", 1)[0].strip()
    if normalized == "image/jpg":
        normalized = "image/jpeg"
    if normalized in _SAFE_IMAGE_MIMES or normalized in {"image/heic", "image/heif"}:
        return normalized

    suffix_map = {
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".webp": "image/webp",
        ".gif": "image/gif",
        ".heic": "image/heic",
        ".heif": "image/heif",
        ".bmp": "image/bmp",
        ".dib": "image/bmp",
        ".tif": "image/tiff",
        ".tiff": "image/tiff",
    }
    mapped = suffix_map.get(suffix)
    if mapped:
        return mapped
    if normalized.startswith("image/"):
        return normalized
    return "image/png"


def _image_to_png_bytes(path: Path) -> bytes:
    from PIL import Image

    with Image.open(path) as image:
        converted = image.convert("RGBA" if "A" in image.getbands() else "RGB")
        buffer = io.BytesIO()
        converted.save(buffer, format="PNG", optimize=True)
        return buffer.getvalue()


def image_to_data_url_with_meta(path: str, mime: str, *, locale: str = "zh-CN") -> tuple[str, str | None]:
    """
    Returns (data_url, warning). For HEIC, fallback to original HEIC payload
    when local conversion is unavailable, so capable gateways can still consume it.
    """
    file_path = Path(path)
    suffix = file_path.suffix.lower()
    raw: bytes
    out_mime = _normalize_image_mime(mime, suffix)
    warning: str | None = None

    is_heic = suffix in {".heic", ".heif"} or out_mime in {"image/heic", "image/heif"}
    if is_heic:
        try:
            raw = _heic_to_jpeg_bytes(file_path)
            out_mime = "image/jpeg"
        except Exception:
            raw = file_path.read_bytes()
            out_mime = "image/heic"
            warning = _attachment_text(locale, "image.heic_warning")
    else:
        if out_mime not in _SAFE_IMAGE_MIMES:
            original_mime = out_mime
            try:
                raw = _image_to_png_bytes(file_path)
                out_mime = "image/png"
                warning = _attachment_text(locale, "image.converted_warning", mime=original_mime)
            except Exception as exc:
                raise RuntimeError(_attachment_text(locale, "image.unsupported_type", mime=original_mime, error=exc)) from exc
        else:
            raw = file_path.read_bytes()

    if not raw:
        raise RuntimeError(_attachment_text(locale, "image.empty"))

    encoded = base64.b64encode(raw).decode("ascii")
    return f"data:{out_mime};base64,{encoded}", warning


def image_to_data_url(path: str, mime: str) -> str:
    data_url, _ = image_to_data_url_with_meta(path, mime)
    return data_url


def summarize_file_payload(path: str, max_bytes: int = 768, max_text_chars: int = 1200, *, locale: str = "zh-CN") -> str:
    file_path = Path(path)
    raw = file_path.read_bytes()
    head = raw[:max_bytes]

    if not head:
        return _attachment_text(locale, "file.empty")

    text_bytes = b"\n\r\t\b\f" + bytes(range(32, 127))
    non_text = sum(1 for b in head if b not in text_bytes)
    is_binary = b"\x00" in head or (non_text / len(head)) > 0.30

    if not is_binary:
        text = head.decode("utf-8", errors="ignore")
        text = text[:max_text_chars]
        return _attachment_text(locale, "file.text_preview", size=len(raw), text=text)

    hex_preview = " ".join(f"{b:02x}" for b in head[:128])
    return (
        _attachment_text(
            locale,
            "file.binary_preview",
            size=len(raw),
            preview_size=min(len(head), 128),
            hex_preview=hex_preview,
        )
    )
